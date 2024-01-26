from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
import func
from time import sleep
import openpyxl
import os

driver = webdriver.Chrome()
driver.maximize_window()
driver.get("https://www.mercadolivre.com/jms/mlb/lgz/msl/login/H4sIAAAAAAAEAzWNSw6DMAxE7-I1gn2WvUhkgoGoDkGOIVQVd69TieV83swXOC9x8_rZCRzQtXMMUaGDnVHnLMnHyYLEZpWo9MixVVAwkZIUcN82tND0IoPa1IxcyEp46OpnztW8_5d5sXi6jNuQfaXxjNTSh1iyiVV1L24Yaq19Igk4ZY6nUB9y6kcZ4O4MKOpVMLzBqRx0_wBJfknyzgAAAA/user")
sleep(3)

inputs = openpyxl.load_workbook('input.xlsx')
sheet = inputs.active
site_name = sheet.cell(row = 2, column = 1).value
email = sheet.cell(row = 2, column = 2).value
password = sheet.cell(row = 2, column = 3).value
num_of_products = [cell.value for cell in inputs['Sheet1']['D']]
num_of_products = len([cell.value for cell in inputs['Sheet1']['D']]) - 1

print("Email: ", email)
print("Password: ", password)
print("Total number of products: ", num_of_products)

if not os.path.exists('screenshot'):
  os.makedirs('screenshot')

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Website'
ws['B1'] = 'Piece Code'
ws['C1'] = 'Product Name'
ws['D1'] = 'Reference(R$)'
ws['E1'] = 'Description'
ws['F1'] = 'Price(R$)'
ws['G1'] = 'Brand'
ws['H1'] = 'Seller'
ws['I1'] = 'Link'
ws['J1'] = 'Screenshot'
wb.save('result.xlsx')

while True:
    response = input('Are you ready? (y/n): ')
    if response =='y':
        break
    sleep(1)

match_num = 0
for i in range(num_of_products):
  sleep(0.3)
  driver.get("https://www.mercadolivre.com.br/")
  func.wait_url(driver, 'https://www.mercadolivre.com.br/')
  print("Finding {}th product...".format(i + 1))

  sheet = inputs.active
  piece_code = sheet.cell(row = (i + 2), column = 4).value
  product_name = sheet.cell(row = (i + 2), column = 5).value
  reference_price = sheet.cell(row = (i + 2), column = 6).value
  func.find_element(driver, By.ID, 'cb1-edit').send_keys(piece_code)
  func.find_element(driver, By.CSS_SELECTOR, 'button[class="nav-search-btn"]').click()
  link = driver.current_url
  products = driver.find_element(By.CSS_SELECTOR, 'ol[class="ui-search-layout ui-search-layout--stack"]').find_elements(By.TAG_NAME, 'li')
  flag = False
  for product in products:
    try:
      product.find_element(By.TAG_NAME, 'a').click()
      flag = True
      break
    except:
      continue
  if flag == False:
    print("No result.")
    continue
  sleep(0.1)
  try:
    price_int = driver.find_element(By.CSS_SELECTOR, 'span[class="andes-money-amount__fraction"]').text
  except:
    continue
  price_int = price_int.replace(",", "").replace(".", "")
  price_int = int(price_int)
  try:
    price_fraction = int(driver.find_element(By.CSS_SELECTOR, 'span[class="andes-money-amount__cents andes-money-amount__cents--superscript-36"]').text)
  except:
    price_fraction = 0
  price = float(price_int + price_fraction/100)
  if price > reference_price:
    print("Expensive(R$ {})".format(price))
    continue
  try:
    description = driver.find_element(By.CSS_SELECTOR, 'div[class="ui-pdp-header__title-container"]').find_element(By.TAG_NAME, 'h1').text
  except:
    description = "Not found"
  try:
    brand = driver.find_element(By.CSS_SELECTOR, 'table[class="andes-table"]').find_elements(By.TAG_NAME, 'tr')[0].find_element(By.TAG_NAME, 'td').find_element(By.TAG_NAME, 'span').text
  except:
    brand = "Not found"
  try:
    seller = driver.find_element(By.CSS_SELECTOR, 'div[class="ui-pdp-seller__header__title"]').find_element(By.TAG_NAME, 'span').text
  except:
    seller = "Not found"
  image_name = '{}_{}.png'.format(site_name, piece_code)
  try:
    driver.find_element(By.CSS_SELECTOR, 'div[class="ui-pdp-container ui-pdp-container--pdp"]').screenshot('screenshot/{}'.format(image_name))
  except:
    image_name = "Not found"
  
  match_num += 1
  workbook = openpyxl.load_workbook('result.xlsx')
  sheet = workbook['Sheet']
  sheet[f'A{match_num + 1}'] = site_name
  sheet[f'B{match_num + 1}'] = piece_code
  sheet[f'C{match_num + 1}'] = product_name
  sheet[f'D{match_num + 1}'] = reference_price
  sheet[f'E{match_num + 1}'] = description
  sheet[f'F{match_num + 1}'] = price
  sheet[f'G{match_num + 1}'] = brand
  sheet[f'H{match_num + 1}'] = seller
  sheet[f'I{match_num + 1}'] = link
  sheet[f'J{match_num + 1}'] = image_name
  workbook.save('result.xlsx')
  
  print("Successful!")

print("{} products found!".format(match_num))
print("Finishing...")
driver.quit()