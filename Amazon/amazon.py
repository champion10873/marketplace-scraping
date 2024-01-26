from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
import func
from time import sleep
import openpyxl
import os

driver = webdriver.Chrome()
driver.maximize_window()

inputs = openpyxl.load_workbook('entrada.xlsx')
sheet = inputs.active
site_name = sheet.cell(row = 2, column = 1).value
email = sheet.cell(row = 2, column = 2).value
password = sheet.cell(row = 2, column = 3).value
num_of_products = len([cell.value for cell in inputs['Sheet1']['D']]) - 1

print("E-mail: ", email)
print("Senha: ", password)
print("Número total de produtos: ", num_of_products)

if not os.path.exists('captura de tela'):
  os.makedirs('captura de tela')

wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 'Site'
ws['B1'] = 'Código da peça'
ws['C1'] = 'Nome do produto'
ws['D1'] = 'Referência(R$)'
ws['E1'] = 'Descrição'
ws['F1'] = 'Preço(R$)'
ws['G1'] = 'Marca'
ws['H1'] = 'Vendedor'
ws['I1'] = 'Entregar'
ws['J1'] = 'Link'
ws['K1'] = 'Captura de tela'
wb.save('resultado.xlsx')

driver.get("https://www.amazon.com.br/ap/signin?openid.pape.max_auth_age=0&openid.return_to=https%3A%2F%2Fwww.amazon.com.br%2F%3Fref_%3Dnav_custrec_signin&openid.identity=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.assoc_handle=brflex&openid.mode=checkid_setup&openid.claimed_id=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0%2Fidentifier_select&openid.ns=http%3A%2F%2Fspecs.openid.net%2Fauth%2F2.0")
func.find_element(driver, By.ID, "ap_email").send_keys(email)
func.find_element(driver, By.ID, "continue").click()
func.find_element(driver, By.ID, "ap_password").send_keys(password)
func.find_element(driver, By.ID, "signInSubmit").click()

while True:
  response = input("Did you pass the character? (y/n): ")
  if response == 'y':
    break

match_num = 0
for i in range(num_of_products):
  sleep(0.3)
  print("Encontrar o produto {}...".format(i + 1))
  driver.get('https://www.amazon.com.br/')
  func.wait_url(driver, 'https://www.amazon.com.br/')

  sheet = inputs.active
  piece_code = sheet.cell(row = (i + 2), column = 4).value
  search_name = sheet.cell(row = (i + 2), column = 5).value
  reference_price = sheet.cell(row = (i + 2), column = 6).value
  func.find_element(driver, By.ID, 'twotabsearchtextbox').send_keys(piece_code)
  func.find_element(driver, By.ID, 'nav-search-submit-button').click()
  sleep(0.2)
  
  products = driver.find_element(By.ID, 'search').find_element(By.CSS_SELECTOR, 'div[class="s-main-slot s-result-list s-search-results sg-row"]').find_elements(By.CLASS_NAME, 's-result-item')
  products_num = len(products)
  # print(len(products)) # --------------------------

  is_result = products[0].find_element(By.CSS_SELECTOR, 'div[class="s-no-outline"]').find_elements(By.TAG_NAME, 'span')[0].text
  # print(is_result) # --------------------------
  if is_result.find("Nenhum resultado") != -1:
    print("Nenhum resultado.")
    continue
    
  for i in range(products_num):
    sleep(0.1)
    products = driver.find_element(By.ID, 'search').find_element(By.CSS_SELECTOR, 'div[class="s-main-slot s-result-list s-search-results sg-row"]').find_elements(By.CLASS_NAME, 's-result-item')
    try:
      price_int = products[i].find_element(By.CSS_SELECTOR, 'span[class="a-price-whole"]').text
      # print(price_int) # --------------------------
    except:
      # print("Não é possível encontrar o preço.") # --------------------------
      continue
    price_int = price_int.replace(",", "").replace(".", "")
    price_int = int(price_int)
    try:
      price_fraction = int(products[i].find_element(By.CSS_SELECTOR, 'span[class="a-price-fraction"]').text)
    except:
      price_fraction = 0
    price = float(price_int + price_fraction/100)
    # print(price) # --------------------------
    if price > reference_price:
      # print("Tão caro.") # --------------------------
      continue
    product_name = products[i].find_element(By.CSS_SELECTOR, 'div[data-cy="title-recipe"]').text
    # print(product_name) # --------------------------
    if func.evaluate_similarity(search_name, product_name) == False:
      # print("Produto diferente.") # --------------------------
      continue
    try:
      products[i].find_element(By.TAG_NAME, 'img').click()
    except:
      # print("Indisponível para clicar.")
      continue
    print("Produto encontrado.")
    sleep(0.2)
    try:
      brand = driver.find_element(By.CSS_SELECTOR, 'span[class="a-size-base po-break-word"]').text
    except:
      brand = "Não encontrado"
    try:
      seller = driver.find_elements(By.CSS_SELECTOR, 'span[class="a-size-small offer-display-feature-text-message"]')[0].text
    except:
      seller = "Não encontrado"
    try:
      deliver = driver.find_elements(By.CSS_SELECTOR, 'span[class="a-size-small offer-display-feature-text-message"]')[1].text
    except:
      deliver = "Não encontrado"
    link = driver.current_url
    match_num += 1
    image_name = '{}_{}.png'.format(site_name, match_num)
    try:
      driver.find_element(By.ID, 'dp').screenshot('captura de tela/{}'.format(image_name))
    except:
      image_name = "Não encontrado"
  
    workbook = openpyxl.load_workbook('resultado.xlsx')
    sheet = workbook['Sheet']
    sheet[f'A{match_num + 1}'] = site_name
    sheet[f'B{match_num + 1}'] = piece_code
    sheet[f'C{match_num + 1}'] = search_name
    sheet[f'D{match_num + 1}'] = reference_price
    sheet[f'E{match_num + 1}'] = product_name
    sheet[f'F{match_num + 1}'] = price
    sheet[f'G{match_num + 1}'] = brand
    sheet[f'H{match_num + 1}'] = seller
    sheet[f'I{match_num + 1}'] = deliver
    sheet[f'J{match_num + 1}'] = link
    sheet[f'K{match_num + 1}'] = image_name
    workbook.save('resultado.xlsx')
    print("Successful!")
    driver.back()

print("{} produtos encontrados!".format(match_num))
print("Desistindo...")
driver.quit()